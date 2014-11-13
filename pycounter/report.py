"""COUNTER journal and book reports and associated functions"""

from __future__ import absolute_import

import logging
import re
import warnings

import pyisbn
import six
from six.moves import range

from pycounter import _csvhelper


class UnknownReportTypeError(Exception):
    pass


class CounterReport(object):
    def __init__(self):
        self.pubs = []
        self.year = None
        self.report_type = None
        self.report_version = 0

    def __str__(self):
        return "CounterReport %s version %s for %s" % (self.report_type,
                                                       self.report_version,
                                                       self.year)

    def __iter__(self):
        return iter(self.pubs)


class CounterPublication(object):
    def __init__(self, line=None):
        if line is not None:
            self.title = line[0]
            self.publisher = line[1]
            self.platform = line[2]
            self.issn = line[3].strip()
            self.eissn = line[4].strip()
            self.isbn = None
            self.monthdata = [format_stat(x) for x in line[5:]]
            while len(self.monthdata) < 12:
                self.monthdata.append(None)
            logging.debug("monthdata: %s", self.monthdata)

    def __str__(self):
        return """<CounterPublication %s, publisher %s,
        platform %s>""" % (self.title, self.publisher, self.platform)


class CounterBook(object):
    def __init__(self, line=None):
        if line is not None:
            self.title = line[0]
            self.publisher = line[1]
            self.platform = line[2]
            self.isbn = line[3].strip().replace('-', '')
            if len(self.isbn) == 10:
                self.isbn = pyisbn.convert(self.isbn)
            self.issn = line[4].strip()
            self.eissn = None
            self.monthdata = [format_stat(x) for x in line[5:]]
            while len(self.monthdata) < 12:
                self.monthdata.append(None)
            logging.debug("monthdata: %s", self.monthdata)

    def __str__(self):
        return """<CounterPublication %s (ISBN: %s), publisher %s,
        platform %s>""" % (self.title, self.isbn, self.publisher,
                           self.platform)


def format_stat(stat):
    stat = stat.replace(',', '')
    try:
        return int(stat)
    except ValueError:
        return None


def parse(filename):
    """Parse a COUNTER file, first attempting to determine type"""
    if filename.endswith('.tsv'):
        # Horrible filename-based hack; in future examine contents of file here
        return parse_separated(filename, '\t')
    if filename.endswith('.xlsx'):
        return parse_xlsx(filename)
    # fallback to old assume-csv behavior
    return parse_separated(filename, ',')


def parse_xlsx(filename):
    from openpyxl import load_workbook
    workbook = load_workbook(filename=filename)
    worksheet = workbook.get_sheet_by_name(workbook.get_sheet_names()[0])
    row_it = worksheet.iter_rows()
    split_row_list = ([cell.value if cell.value is not None else ""
                       for cell in row] for row in row_it)

    return parse_generic(split_row_list)


def parse_separated(filename, delimiter):
    """Open COUNTER CSV/TSV report with given filename and delimiter
    and parse into a CounterReport object"""
    with _csvhelper.UnicodeReader(filename,
                                  delimiter=delimiter) as report_reader:
        return parse_generic(report_reader)


def parse_generic(report_reader):
    """Takes an iterator of COUNTER report rows and
    returns a CounterReport object"""
    report = CounterReport()

    line1 = six.next(report_reader)

    rt_match = re.match(
        r'.*(Journal|Book|Database) Report (\d) ?\(R(\d)\)',
        line1[0])
    if rt_match:
        report.report_type = (rt_match.group(1)[0].capitalize() + 'R' +
                              rt_match.group(2))
        report.report_version = int(rt_match.group(3))

    for _ in range(3):
        six.next(report_reader)
    if report.report_version == 4:
        # COUNTER 4 has 3 more lines of introduction
        for _ in range(3):
            six.next(report_reader)
    header = six.next(report_reader)
    first_date_col = 10 if report.report_version == 4 else 5
    if report.report_type in ('BR1', 'BR2') and report.report_version == 4:
        first_date_col = 8

    report.year = int(header[first_date_col].split('-')[1])
    if report.year < 100:
        report.year += 2000

    if report.report_version == 4:
        countable_header = header[0:8]
        for col in header[8:]:
            if col:
                countable_header.append(col)
        last_col = len(countable_header)
    else:
        for last_col, v in enumerate(header):
            if 'YTD' in v:
                break
    six.next(report_reader)
    for line in report_reader:
        if not line:
            continue
        if report.report_version == 4:
            if report.report_type == 'JR1':
                line = line[0:3] + line[5:7] + line[10:last_col]
            elif report.report_type in ('BR1', 'BR2'):
                line = line[0:3] + line[5:7] + line[8:last_col]
        else:
            line = line[0:last_col]
        logging.debug(line)
        if report.report_type:
            if report.report_type.startswith('JR'):
                report.pubs.append(CounterPublication(line))
            elif report.report_type.startswith('BR'):
                report.pubs.append(CounterBook(line))
            else:
                raise UnknownReportTypeError(report.report_type)

    return report


def parse_csv(filename):
    warnings.warn(".parse_csv is deprecated; use parse_separated",
                  DeprecationWarning)
    return parse_separated(filename, ",")


def parse_tsv(filename):
    warnings.warn(".parse_tsv is deprecated; use parse_separated",
                  DeprecationWarning)
    return parse_separated(filename, "\t")